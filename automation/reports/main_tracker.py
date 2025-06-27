from msilib.schema import Font
import re
import sys
import os
import time
import pdb
from tkinter import messagebox
import zipfile
import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
import logging
from openpyxl.styles import PatternFill,Font
import linecache
import warnings
from datetime import datetime,timedelta

import numpy as np
warnings.filterwarnings('ignore', message='Workbook contains no default style, apply openpyxl\'s default')
warnings.filterwarnings('ignore', message='Conditional Formatting extension is not supported and will be removed')


current_dir = str(os.path.dirname(os.path.realpath(__file__)))


template_path = os.path.join(current_dir, 'Template')
output_path = os.path.join(current_dir, 'output')

for filename in os.listdir(output_path):
    if filename.endswith('.xlsx'):
        os.remove(os.path.join(output_path, filename))

template_file = os.path.join(template_path, f'Template.xlsx')


def tracker_(file_1,file_2,file_3):

    column_names = [
        "A", "B", "C", "D", "E", 
        "F", "G", "H", "I", 
        "J", "K", "L", "M", "N", 

        "O", "P", "Q", "R", "S", "T", 
        "U", 

        "V", "W", "X", "Y", "Z", 
        "AA", "AB", "AC", "AD", "AE", 
        "AF",

        "AG", "AH", "AI", "AJ", "AK", 
        "AL", "AM", "AN", "AO", "AP", 
        "AQ"
    ]

    df_1 = pd.read_excel(file_1, sheet_name='Result',usecols=["A", "B", "C"])
    df = pd.read_excel(file_2, skiprows=2, header=None)
    df.columns = column_names
    df_2 = pd.read_excel(file_3, engine='openpyxl', sheet_name='sheet1' , skiprows=3, header=None)
    df_selected = df_2.iloc[:, [0, 1, 2, 31, 41]]
    df_selected.columns = ["A", "B", "C", "D", "E"]

    def eliminar_hasta_guion(texto):
        texto = str(texto)
        partes = texto.split('_', 1)
        if len(partes) > 1:
            return '_' + partes[1]
        return texto

    df_cleaned = df.dropna(subset=['AM'])
    df_cleaned['AM'] = pd.to_datetime(df_cleaned['AM'])
    yesterday = (datetime.now() - timedelta(days=1)).date()
    df_filtered = df_cleaned[df_cleaned['AM'].dt.date == yesterday]
    df_filtered_reset = df_filtered.reset_index(drop=True)
    df_id = df_filtered_reset['D']
    df_id = df_id.to_frame() 
    df['lookup'] = df['D'].apply(lambda x:  df_id.loc[df_id['D'] == x,  'D'].iloc[0] if x in  df_id['D'].values else "N/A")
    df = df[df['lookup'] != 'N/A']
    df['short'] = df['D'].apply(eliminar_hasta_guion)
    df_selected['short_b'] = df_selected['B'].apply(eliminar_hasta_guion)
    df['extra1'] = df['short'].apply(lambda x:  df_selected.loc[df_selected['short_b'] == x,  'E'].iloc[0] if x in  df_selected['short_b'].values else "")
    df['B_selected'] = df['short'].apply(lambda x:  df_selected.loc[df_selected['short_b'] == x,  'B'].iloc[0] if x in  df_selected['short_b'].values else "")
    df['C_selected'] = df['short'].apply(lambda x:  df_selected.loc[df_selected['short_b'] == x,  'C'].iloc[0] if x in  df_selected['short_b'].values else "")
    df['D_selected'] = df['short'].apply(lambda x:  df_selected.loc[df_selected['short_b'] == x,  'D'].iloc[0] if x in  df_selected['short_b'].values else "")
    df['E_selected'] = df['B_selected'].apply(lambda x:  df_1.loc[df_1['DU ID'] == x,  'Actual Task Close Time'].iloc[0] if x in  df_1['DU ID'].values else "")
    df['F_selected'] = df['B_selected'].apply(lambda x:  df_1.loc[df_1['DU ID'] == x,  'Historical Approver'].iloc[0] if x in  df_1['DU ID'].values else "")

    df_filtered_reset['short_reset'] = df_filtered_reset['D'].apply(eliminar_hasta_guion)
    unique_site_ids = df_filtered_reset['short_reset'].unique().tolist()
    expected_groups = {"A", "B", "C", "D"}
    valid_dfs = []

    for site_id in unique_site_ids:
        filtered_df = df[df['short'] == site_id] 
        filtered_df = filtered_df[filtered_df['AL'] != 'Reviewing']
        filtered_df = filtered_df[filtered_df['AL'].notna() & (filtered_df['AL'] != '')]
        filtered_df['AM'] = pd.to_datetime(filtered_df['AM'], errors='coerce')
        groups_in_filtered_df = set(filtered_df['AJ'].unique())

        if expected_groups.issubset(groups_in_filtered_df):
            updated_groups_df = pd.DataFrame()
            for group in expected_groups:
                group_df = filtered_df[filtered_df['AJ'] == group]
                if not group_df.empty:
                    latest_record = group_df.loc[group_df['AM'].idxmax()] 
                    updated_groups_df = updated_groups_df.append(latest_record, ignore_index=True)
            if len(updated_groups_df) == 4:
                valid_dfs.append(updated_groups_df)
        else:
            continue

    dfs_combinados = []
    for valid_df in valid_dfs:
        columnas_seleccionadas = ['B','D','AB','AJ','AK','AL','AM','AN','extra1','B_selected','C_selected','D_selected','E_selected','F_selected']
        df_seleccionado = valid_df[columnas_seleccionadas]
        df_seleccionado['AO'] = datetime.now().strftime('%m/%d/%Y')
        df_seleccionado['AN'] = df_seleccionado.apply(lambda row: 'Approved' if row['AL'] == 'Approved' else (row['AN'].split('\n') if isinstance(row['AN'], str) else 'vacio'), axis=1)
        df_exploded = df_seleccionado.explode('AN', ignore_index=True)
        dfs_combinados.append(df_exploded)

    df_general = pd.concat(dfs_combinados, ignore_index=True)
    df_general = df_general[df_general['AN'].notna() & (df_general['AN'].str.strip() != '')]
    df_general['AP'] = '=Week_Date_QC_Closed'
    df_general['AQ'] = '=Week_SF_Delivered_V1'
    df_general['AR'] = '=Week_feedback_by_customer_SF_V1'
    fill_teamplate(template_file,df_general,output_path)

 

def fill_teamplate(template_file, df_general, output_path):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    for i, row in df_general.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)
    wb.save(os.path.join(output_path, 'Template.xlsx'))
    

def logic():
    xlsx_path = "file.xlsx"
    xlsm_path = "file.xlsm"

    try:
        if os.path.exists(xlsx_path):
            tracker_(
                "file_.xlsx",
                xlsx_path,
                "file_d.xlsx"
            )
            print("Función con .xlsx DONE")
        else:
            tracker_(
                "file_.xlsx",
                xlsm_path,
                "file_d.xlsx"
            )
            print("Función con .xlsm DONE")

    except Exception as e:
        print(f"Error: {e}")

#logic()